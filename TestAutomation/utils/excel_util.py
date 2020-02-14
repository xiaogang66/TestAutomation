"""使用openpyxl封装的excel操作工具类"""

from openpyxl import load_workbook

class ExcelUtil(object):

    def __init__(self,filePath,sheetName="Sheet1"):
        self.filePath = filePath
        self.workboot = load_workbook(filePath)
        self.sheet = self.workboot[sheetName]

    def get_all_dict_data(self):
        """读取表格数据转换成字典的列表格式显示"""
        keys = self.getRowValues(1)
        rowNum = self.sheet.max_row   # 获取总行数
        colNum = self.sheet.max_column   # 获取总列数
        if rowNum < 1:
            print("总行数小于 1")
        else:
            result = []
            j = 2
            for i in range(rowNum - 1):
                s = {}
                # 从第二行取对应 values 值
                values = self.getRowValues(j)
                for x in range(colNum):
                    s[keys[x]] = values[x]
                result.append(s)
                j += 1
        return result

    def get_data(self,cellNo):
        """读取指定单元格的数据"""
        return self.sheet[cellNo].value

    def set_data(self,cellNo,value):
        """修改指定单元格的数据"""
        self.sheet[cellNo] = value
        self.workboot.save(self.filePath)

    def create_excel(self,cellList):
        for cell in cellList:
            self.set_data(cell[0],cell[1])
        self.workboot.save(self.filePath)

    def getRowValues(self,row):
        """获取整行的所有值，传入行数"""
        columns = self.sheet.max_column
        rowdata=[]
        for i in range(1,columns+1):
            cellvalue = self.sheet.cell(row=row,column=i).value
            rowdata.append(cellvalue)
        return rowdata

if __name__ == "__main__":
    filepath = "D:\\test.xlsx"
    sheetName = "Sheet1"
    excelutil = ExcelUtil(filepath, sheetName)
    # 创建
    cellList = [('A1','姓名'),('B1','性别'),('C1','年龄'),('D1','地址'),('A2','张三'),('B2','男'),('C2','25'),('D2','江苏南京'),('A3','李四'),('B3','女'),('C3','25'),('D3','江苏苏州'),('A4','李四2'),('B4','女2'),('C4','252'),('D4','江苏苏州2')]
    excelutil.create_excel(cellList)
    # 获取
    data1 = excelutil.get_data("D3")
    print(data1)
    # 设置
    excelutil.set_data('D3','江苏徐州')
    # 获取所有
    dict = excelutil.get_all_dict_data()
    print(dict)
